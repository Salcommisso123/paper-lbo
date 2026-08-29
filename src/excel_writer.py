"""
excel_writer.py — writes the LBO as a LIVE Excel model, not a printout.

WHY THIS IS FORMULA-DRIVEN
--------------------------
The first version of this file pasted the engine's numbers into cells. Every figure was
correct, and the workbook was still the wrong artifact: change the entry multiple and
nothing moved. A paper LBO exists to be flexed in front of someone, and "it's computed
in Python" is a good engineering answer to a question nobody asked.

So the engine's numbers are no longer written at all. The ASSUMPTIONS are written, and
everything downstream is an Excel formula. Open the workbook, change 5.0x to 6.0x, and
the debt schedule, the three statements and the returns all move. The engine and the
workbook now agree because they implement the same arithmetic, which also means the
workbook independently checks the engine — tests/test_excel_recalc.py recalculates it
and compares.

COLOUR CONVENTION (the standard one: blue input, black formula, green cross-sheet)
---------------------------------------------------------------------------------
Colour is derived from the cell's own content in `_put`, never passed in by hand — so a
cell cannot claim to be a formula while holding a hardcode. That is the property the
convention exists to give a reader, and hand-applied colours lose it on the first edit.

NO CIRCULAR REFERENCES
----------------------
Interest is charged on BEGINNING balances, so the chain is a DAG:
    debt beginning -> interest -> net income -> free cash flow -> sweep -> debt ending
Average-balance interest would close that loop and force iterative calculation on, which
is a footgun in a workbook someone else opens.
"""

from __future__ import annotations

from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from edgar import CompanyFinancials, non_cash_nwc
from lbo_engine import LBOAssumptions, LBOResult

NAVY = "1F3864"
WHITE = "FFFFFF"
GREY = "808080"

# The convention. Blue = you may type here; black = calculated here; green = it came
# from another sheet.
INPUT_BLUE = "0000FF"
FORMULA_BLACK = "000000"
LINK_GREEN = "008000"

MONEY = "#,##0;(#,##0)"
MULT = '0.00"x"'
PCT = "0.0%"
PCT2 = "0.00%"

TITLE_FONT = Font(bold=True, size=14, color=NAVY)
SUBTITLE_FONT = Font(italic=True, size=9, color=GREY)
HEADER_FONT = Font(color=WHITE, bold=True, size=11)
HEADER_FILL = PatternFill(fill_type="solid", fgColor=NAVY)
SECTION_FONT = Font(bold=True, size=10, color=NAVY)
TOTAL_BORDER = Border(top=Side(style="thin"), bottom=Side(style="double"))
CHECK_OK = PatternFill(fill_type="solid", fgColor="C6EFCE")


def _put(ws: Worksheet, ref: str, value, fmt: str = MONEY, bold: bool = False,
         indent: int = 0):
    """
    Write a cell and colour it by what it actually contains:
      starts with '=' and mentions another sheet  -> green
      starts with '='                             -> black
      anything else (a hardcode)                  -> blue

    Deriving the colour instead of accepting it as an argument is the point: the
    convention stays true no matter how the sheet is later edited.
    """
    cell = ws[ref]
    cell.value = value
    if isinstance(value, str) and value.startswith("="):
        colour = LINK_GREEN if "!" in value else FORMULA_BLACK
        cell.number_format = fmt
    elif isinstance(value, (int, float)):
        colour = INPUT_BLUE
        cell.number_format = fmt
    else:
        colour = FORMULA_BLACK
    cell.font = Font(bold=bold, color=colour)
    if indent:
        cell.alignment = Alignment(indent=indent)
    return cell


def _label(ws: Worksheet, ref: str, text: str, bold: bool = False, indent: int = 0):
    cell = ws[ref]
    cell.value = text
    cell.font = Font(bold=bold)
    if indent:
        cell.alignment = Alignment(indent=indent)
    return cell


def _section(ws: Worksheet, row: int, text: str):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT


def _title(ws: Worksheet, text: str, subtitle: str = ""):
    ws["A1"] = text
    ws["A1"].font = TITLE_FONT
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = SUBTITLE_FONT


def _year_header(ws: Worksheet, row: int, years: int, first_col: int = 3,
                 opening_label: Optional[str] = "At close"):
    """Column headers: an optional opening column, then Year 1..N."""
    if opening_label:
        c = ws.cell(row=row, column=first_col - 1, value=opening_label)
        c.font, c.fill = HEADER_FONT, HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    for i in range(years):
        c = ws.cell(row=row, column=first_col + i, value=f"Year {i + 1}")
        c.font, c.fill = HEADER_FONT, HEADER_FILL
        c.alignment = Alignment(horizontal="center")


def _widths(ws: Worksheet, first: int = 34, rest: int = 15, n: int = 8):
    ws.column_dimensions["A"].width = first
    for i in range(2, 2 + n):
        ws.column_dimensions[get_column_letter(i)].width = rest


# --------------------------------------------------------------------------------
# Assumptions — the only sheet with hardcodes. Everything else points back here.
# --------------------------------------------------------------------------------
A = {  # row anchors on the Assumptions sheet, referenced by every other sheet
    "revenue": 5, "ebitda": 6, "margin": 7, "ppe": 8, "nwc": 9, "cash": 10,
    "entry": 12, "exit": 13, "leverage": 14, "hold": 15, "growth": 16,
    "da": 17, "capex": 18, "nwc_pct": 19, "tax": 20,
    "txn_fee": 21, "fin_fee": 22, "sweep": 23, "rollover": 24,
    "sr_pct": 27, "sr_rate": 28, "sr_amort": 29, "sub_rate": 30,
}
ASM = "Assumptions"


def _a(key: str) -> str:
    """Absolute reference to an assumption cell, e.g. Assumptions!$B$12."""
    return f"{ASM}!$B${A[key]}"


def _write_assumptions(ws: Worksheet, company: CompanyFinancials, ltm_revenue: float,
                       ltm_ebitda: float, a: LBOAssumptions):
    _title(ws, f"{company.company_name} ({company.ticker}) — LBO Assumptions",
           f"CIK {company.cik} · SEC EDGAR. Blue cells are inputs — change them and the "
           f"whole model recalculates.")
    year = company.latest_complete_year()
    opening_ppe = (year.ppe_net if year and year.ppe_net else 0.0)
    opening_nwc = (non_cash_nwc(year) if year else None) or 0.0

    _section(ws, 4, "LTM financials (from SEC 10-K filings)")
    _label(ws, "A5", "LTM revenue", indent=1);        _put(ws, "B5", ltm_revenue)
    _label(ws, "A6", "LTM EBITDA", indent=1);         _put(ws, "B6", ltm_ebitda)
    _label(ws, "A7", "EBITDA margin", indent=1);      _put(ws, "B7", "=B6/B5", PCT)
    _label(ws, "A8", "Opening PP&E, net", indent=1);  _put(ws, "B8", opening_ppe)
    _label(ws, "A9", "Opening non-cash NWC", indent=1); _put(ws, "B9", opening_nwc)
    # Cash-free/debt-free entry: the sponsor inherits no cash and builds it from retained
    # FCF. An input rather than a hard zero, so a cash-funded deal can be modelled.
    _label(ws, "A10", "Opening cash at close", indent=1); _put(ws, "B10", 0.0)

    _section(ws, 11, "Deal assumptions")
    rows = [
        ("entry", "Entry EV / EBITDA", a.entry_ev_multiple, MULT),
        ("exit", "Exit EV / EBITDA", a.exit_ev_multiple, MULT),
        ("leverage", "Total leverage (x EBITDA)", a.leverage_multiple, MULT),
        ("hold", "Hold period (years)", a.hold_period_years, "0"),
        ("growth", "Revenue growth (annual)", a.revenue_growth_rate, PCT),
        ("da", "D&A (% of revenue)", a.da_pct_revenue, PCT),
        ("capex", "Capex (% of revenue)", a.capex_pct_revenue, PCT),
        ("nwc_pct", "NWC (% of revenue change)", a.nwc_pct_of_revenue_change, PCT),
        ("tax", "Cash tax rate", a.tax_rate, PCT),
        ("txn_fee", "Transaction fees (% of EV)", a.transaction_fee_pct_of_ev, PCT),
        ("fin_fee", "Financing fees (% of debt)", a.financing_fee_pct_of_debt, PCT),
        ("sweep", "Cash sweep (% of FCF)", a.cash_sweep_pct, PCT),
        ("rollover", "Management rollover (% of equity)", a.management_rollover_pct, PCT),
    ]
    for key, text, value, fmt in rows:
        _label(ws, f"A{A[key]}", text, indent=1)
        _put(ws, f"B{A[key]}", value, fmt)

    senior = a.tranches[0]
    sub = a.tranches[1] if len(a.tranches) > 1 else None
    _section(ws, 26, "Debt structure")
    _label(ws, f"A{A['sr_pct']}", f"{senior.name} (% of debt)", indent=1)
    _put(ws, f"B{A['sr_pct']}", senior.pct_of_total_debt, PCT)
    _label(ws, f"A{A['sr_rate']}", f"{senior.name} interest rate", indent=1)
    _put(ws, f"B{A['sr_rate']}", senior.interest_rate, PCT2)
    _label(ws, f"A{A['sr_amort']}", "Mandatory amortisation (% of principal p.a.)", indent=1)
    _put(ws, f"B{A['sr_amort']}", senior.mandatory_amort_pct_of_principal, PCT2)
    _label(ws, f"A{A['sub_rate']}", f"{sub.name if sub else 'Subordinated'} interest rate", indent=1)
    _put(ws, f"B{A['sub_rate']}", sub.interest_rate if sub else 0.0, PCT2)

    ws["A32"] = ("Colour convention: blue = hardcoded input · black = formula on this "
                 "sheet · green = links to another sheet.")
    ws["A32"].font = SUBTITLE_FONT
    _widths(ws, first=38, rest=16)


# --------------------------------------------------------------------------------
SU = "Sources & Uses"
SU_ROWS = {"ev": 5, "txn": 6, "fin": 7, "uses": 8,
           "debt": 11, "roll": 12, "sponsor": 13, "sources": 14, "check": 16}


def _su(key: str) -> str:
    return f"'{SU}'!$B${SU_ROWS[key]}"


def _write_sources_uses(ws: Worksheet):
    _title(ws, "Sources & Uses", "Every figure is a formula. Sponsor equity is the plug.")
    _section(ws, 4, "Uses")
    _label(ws, "A5", "Purchase enterprise value", indent=1)
    _put(ws, "B5", f"={_a('entry')}*{_a('ebitda')}")
    _label(ws, "A6", "Transaction fees", indent=1)
    _put(ws, "B6", f"={_a('txn_fee')}*B5")
    _label(ws, "A7", "Financing fees", indent=1)
    _put(ws, "B7", f"={_a('fin_fee')}*B11")
    _label(ws, "A8", "Total uses", bold=True)
    _put(ws, "B8", "=SUM(B5:B7)", bold=True)
    ws["B8"].border = TOTAL_BORDER

    _section(ws, 10, "Sources")
    _label(ws, "A11", "New debt", indent=1)
    _put(ws, "B11", f"={_a('leverage')}*{_a('ebitda')}")
    _label(ws, "A12", "Management rollover", indent=1)
    _put(ws, "B12", f"={_a('rollover')}*B5")
    _label(ws, "A13", "Sponsor equity (plug)", indent=1)
    _put(ws, "B13", "=B8-B11-B12")
    _label(ws, "A14", "Total sources", bold=True)
    _put(ws, "B14", "=SUM(B11:B13)", bold=True)
    ws["B14"].border = TOTAL_BORDER

    _label(ws, "A16", "Check: sources − uses", bold=True)
    c = _put(ws, "B16", "=B14-B8", bold=True)
    c.fill = CHECK_OK
    _label(ws, "A17", "Equity as % of total uses", indent=1)
    _put(ws, "B17", "=B13/B8", PCT)
    _widths(ws, first=34, rest=18)


# --------------------------------------------------------------------------------
DS = "Debt Schedule"
IS_ = "Income Statement"
CF = "Cash Flow"
BS = "Balance Sheet"


def _write_debt_schedule(ws: Worksheet, years: int):
    """
    Senior is repaid before subordinated. Interest is on beginning balances, which is
    what keeps the workbook free of circular references.
    """
    _title(ws, "Debt Schedule",
           "Mandatory amortisation first, then the sweep, senior before subordinated.")
    _year_header(ws, 4, years, first_col=3, opening_label="At close")
    col = lambda i: get_column_letter(3 + i)          # Year i (0-based) -> column
    prev = lambda i: get_column_letter(2 + i)         # the column to its left

    r = {"sr_beg": 6, "sr_int": 7, "sr_amort": 8, "sr_sweep": 9, "sr_end": 10,
         "sb_beg": 12, "sb_int": 13, "sb_sweep": 14, "sb_end": 15,
         "int_tot": 17, "amort_tot": 18, "avail": 19, "swept_tot": 20, "debt_tot": 21}

    _section(ws, 5, "Senior Term Loan")
    for key, text in (("sr_beg", "Beginning balance"), ("sr_int", "Interest expense"),
                      ("sr_amort", "Mandatory amortisation"), ("sr_sweep", "Cash sweep"),
                      ("sr_end", "Ending balance")):
        _label(ws, f"A{r[key]}", text, indent=1, bold=key.endswith("end"))
    _section(ws, 11, "Subordinated Notes")
    for key, text in (("sb_beg", "Beginning balance"), ("sb_int", "Interest expense"),
                      ("sb_sweep", "Cash sweep"), ("sb_end", "Ending balance")):
        _label(ws, f"A{r[key]}", text, indent=1, bold=key.endswith("end"))
    _section(ws, 16, "Total")
    for key, text in (("int_tot", "Total interest expense"),
                      ("amort_tot", "Total mandatory amortisation"),
                      ("avail", "Cash available for sweep"),
                      ("swept_tot", "Total swept to debt"),
                      ("debt_tot", "Total debt outstanding")):
        _label(ws, f"A{r[key]}", text, indent=1, bold=key == "debt_tot")

    # opening balances at close (column B)
    _put(ws, f"B{r['sr_beg']}", f"={_su('debt')}*{_a('sr_pct')}")
    _put(ws, f"B{r['sb_beg']}", f"={_su('debt')}*(1-{_a('sr_pct')})")
    _put(ws, f"B{r['debt_tot']}", f"=B{r['sr_beg']}+B{r['sb_beg']}", bold=True)

    for i in range(years):
        c, p = col(i), prev(i)
        # Senior
        _put(ws, f"{c}{r['sr_beg']}", f"={p}{r['sr_end'] if i else r['sr_beg']}")
        _put(ws, f"{c}{r['sr_int']}", f"={c}{r['sr_beg']}*{_a('sr_rate')}")
        _put(ws, f"{c}{r['sr_amort']}",
             f"=MIN({_a('sr_amort')}*$B${r['sr_beg']},{c}{r['sr_beg']})")
        _put(ws, f"{c}{r['sr_sweep']}",
             f"=MIN({c}{r['avail']},{c}{r['sr_beg']}-{c}{r['sr_amort']})")
        _put(ws, f"{c}{r['sr_end']}",
             f"={c}{r['sr_beg']}-{c}{r['sr_amort']}-{c}{r['sr_sweep']}", bold=True)
        # Subordinated — takes only what the senior sweep left over
        _put(ws, f"{c}{r['sb_beg']}", f"={p}{r['sb_end'] if i else r['sb_beg']}")
        _put(ws, f"{c}{r['sb_int']}", f"={c}{r['sb_beg']}*{_a('sub_rate')}")
        _put(ws, f"{c}{r['sb_sweep']}",
             f"=MIN({c}{r['avail']}-{c}{r['sr_sweep']},{c}{r['sb_beg']})")
        _put(ws, f"{c}{r['sb_end']}", f"={c}{r['sb_beg']}-{c}{r['sb_sweep']}", bold=True)
        # Totals
        _put(ws, f"{c}{r['int_tot']}", f"={c}{r['sr_int']}+{c}{r['sb_int']}")
        _put(ws, f"{c}{r['amort_tot']}", f"={c}{r['sr_amort']}")
        # CF row 10 is free cash flow before debt service. Pointing this one row higher
        # (at the NWC line) silently disables the sweep whenever growth is zero.
        _put(ws, f"{c}{r['avail']}",
             f"=MAX(0,'{CF}'!{c}10-{c}{r['amort_tot']})*{_a('sweep')}")
        _put(ws, f"{c}{r['swept_tot']}", f"={c}{r['sr_sweep']}+{c}{r['sb_sweep']}")
        _put(ws, f"{c}{r['debt_tot']}", f"={c}{r['sr_end']}+{c}{r['sb_end']}", bold=True)

    ws[f"A{r['avail']}"].font = Font(italic=True)
    _widths(ws, first=34, rest=15)
    return r


def _write_income_statement(ws: Worksheet, years: int):
    _title(ws, "Income Statement",
           "Margin is held at the entry margin; D&A is a percentage of revenue.")
    _year_header(ws, 4, years, first_col=3, opening_label="LTM")
    col = lambda i: get_column_letter(3 + i)
    prev = lambda i: get_column_letter(2 + i)
    rows = {"rev": 6, "ebitda": 7, "margin": 8, "da": 9, "ebit": 10,
            "int": 11, "ebt": 12, "tax": 13, "ni": 14}
    for key, text, bold in (("rev", "Revenue", False), ("ebitda", "EBITDA", True),
                            ("margin", "EBITDA margin", False), ("da", "Less: D&A", False),
                            ("ebit", "EBIT", True), ("int", "Less: interest expense", False),
                            ("ebt", "Pre-tax income", True), ("tax", "Less: cash taxes", False),
                            ("ni", "Net income", True)):
        _label(ws, f"A{rows[key]}", text, bold=bold, indent=0 if bold else 1)

    _put(ws, f"B{rows['rev']}", f"={_a('revenue')}")
    _put(ws, f"B{rows['ebitda']}", f"={_a('ebitda')}")
    _put(ws, f"B{rows['margin']}", f"=B{rows['ebitda']}/B{rows['rev']}", PCT)
    for i in range(years):
        c, p = col(i), prev(i)
        _put(ws, f"{c}{rows['rev']}", f"={p}{rows['rev']}*(1+{_a('growth')})")
        _put(ws, f"{c}{rows['ebitda']}", f"={c}{rows['rev']}*{_a('margin')}", bold=True)
        _put(ws, f"{c}{rows['margin']}", f"={c}{rows['ebitda']}/{c}{rows['rev']}", PCT)
        _put(ws, f"{c}{rows['da']}", f"={c}{rows['rev']}*{_a('da')}")
        _put(ws, f"{c}{rows['ebit']}", f"={c}{rows['ebitda']}-{c}{rows['da']}", bold=True)
        _put(ws, f"{c}{rows['int']}", f"='{DS}'!{c}17")
        _put(ws, f"{c}{rows['ebt']}", f"={c}{rows['ebit']}-{c}{rows['int']}", bold=True)
        _put(ws, f"{c}{rows['tax']}", f"=MAX(0,{c}{rows['ebt']}*{_a('tax')})")
        _put(ws, f"{c}{rows['ni']}", f"={c}{rows['ebt']}-{c}{rows['tax']}", bold=True)
    _widths(ws, first=30, rest=15)
    return rows


def _write_cash_flow(ws: Worksheet, years: int):
    _title(ws, "Cash Flow Statement",
           "Free cash flow before debt service feeds the sweep on the Debt Schedule.")
    _year_header(ws, 4, years, first_col=3, opening_label=None)
    col = lambda i: get_column_letter(3 + i)
    prev = lambda i: get_column_letter(2 + i)
    R = {"ni": 6, "da": 7, "capex": 8, "nwc": 9, "fcf": 10,
         "amort": 12, "sweep": 13, "netchg": 14, "begcash": 16, "endcash": 17}
    for key, text, bold in (("ni", "Net income", False), ("da", "Add: D&A (non-cash)", False),
                            ("capex", "Less: capital expenditure", False),
                            ("nwc", "Less: increase in NWC", False),
                            ("fcf", "Free cash flow before debt service", True),
                            ("amort", "Less: mandatory amortisation", False),
                            ("sweep", "Less: cash sweep", False),
                            ("netchg", "Net change in cash", True),
                            ("begcash", "Beginning cash", False),
                            ("endcash", "Ending cash", True)):
        _label(ws, f"A{R[key]}", text, bold=bold, indent=0 if bold else 1)

    for i in range(years):
        c, p = col(i), prev(i)
        _put(ws, f"{c}{R['ni']}", f"='{IS_}'!{c}14")
        _put(ws, f"{c}{R['da']}", f"='{IS_}'!{c}9")
        _put(ws, f"{c}{R['capex']}", f"='{IS_}'!{c}6*{_a('capex')}")
        _put(ws, f"{c}{R['nwc']}", f"=('{IS_}'!{c}6-'{IS_}'!{p}6)*{_a('nwc_pct')}")
        _put(ws, f"{c}{R['fcf']}",
             f"={c}{R['ni']}+{c}{R['da']}-{c}{R['capex']}-{c}{R['nwc']}", bold=True)
        _put(ws, f"{c}{R['amort']}", f"='{DS}'!{c}18")
        _put(ws, f"{c}{R['sweep']}", f"='{DS}'!{c}20")
        _put(ws, f"{c}{R['netchg']}",
             f"={c}{R['fcf']}-{c}{R['amort']}-{c}{R['sweep']}", bold=True)
        _put(ws, f"{c}{R['begcash']}", "=0" if i == 0 else f"={p}{R['endcash']}")
        _put(ws, f"{c}{R['endcash']}",
             f"={c}{R['begcash']}+{c}{R['netchg']}", bold=True)
    _widths(ws, first=36, rest=15)
    return R


def _write_balance_sheet(ws: Worksheet, years: int, cf_rows: dict, is_rows: dict):
    """
    Goodwill is the plug at close, which is what makes the sheet tie: assets must equal
    the total sources that funded them. The check row is the point of the sheet.
    """
    _title(ws, "Balance Sheet",
           "Goodwill is the purchase-accounting plug at close. The check row must be zero.")
    _year_header(ws, 4, years, first_col=3, opening_label="At close")
    col = lambda i: get_column_letter(3 + i)
    prev = lambda i: get_column_letter(2 + i)
    R = {"cash": 6, "nwc": 7, "ppe": 8, "gw": 9, "assets": 10,
         "debt": 13, "equity": 14, "liab_eq": 15, "check": 17}
    for key, text, bold in (("cash", "Cash", False), ("nwc", "Non-cash working capital", False),
                            ("ppe", "PP&E, net", False), ("gw", "Goodwill & intangibles", False),
                            ("assets", "Total assets", True),
                            ("debt", "Total debt", False), ("equity", "Shareholders' equity", False),
                            ("liab_eq", "Total liabilities & equity", True),
                            ("check", "Check: assets − (liabilities + equity)", True)):
        _label(ws, f"A{R[key]}", text, bold=bold, indent=0 if bold else 1)
    _section(ws, 12, "Liabilities & equity")

    # At close
    _put(ws, f"B{R['cash']}", f"={_a('cash')}")
    _put(ws, f"B{R['nwc']}", f"={_a('nwc')}")
    _put(ws, f"B{R['ppe']}", f"={_a('ppe')}")
    _put(ws, f"B{R['gw']}", f"={_su('uses')}-B{R['nwc']}-B{R['ppe']}")
    _put(ws, f"B{R['assets']}", f"=SUM(B{R['cash']}:B{R['gw']})", bold=True)
    _put(ws, f"B{R['debt']}", f"={_su('debt')}")
    _put(ws, f"B{R['equity']}", f"={_su('sponsor')}+{_su('roll')}")
    _put(ws, f"B{R['liab_eq']}", f"=B{R['debt']}+B{R['equity']}", bold=True)
    _put(ws, f"B{R['check']}", f"=B{R['assets']}-B{R['liab_eq']}", bold=True).fill = CHECK_OK

    for i in range(years):
        c, p = col(i), prev(i)
        _put(ws, f"{c}{R['cash']}", f"='{CF}'!{c}{cf_rows['endcash']}")
        _put(ws, f"{c}{R['nwc']}", f"={p}{R['nwc']}+'{CF}'!{c}{cf_rows['nwc']}")
        _put(ws, f"{c}{R['ppe']}",
             f"={p}{R['ppe']}+'{CF}'!{c}{cf_rows['capex']}-'{CF}'!{c}{cf_rows['da']}")
        _put(ws, f"{c}{R['gw']}", f"={p}{R['gw']}")
        _put(ws, f"{c}{R['assets']}", f"=SUM({c}{R['cash']}:{c}{R['gw']})", bold=True)
        _put(ws, f"{c}{R['debt']}", f"='{DS}'!{c}21")
        _put(ws, f"{c}{R['equity']}", f"={p}{R['equity']}+'{IS_}'!{c}{is_rows['ni']}")
        _put(ws, f"{c}{R['liab_eq']}", f"={c}{R['debt']}+{c}{R['equity']}", bold=True)
        _put(ws, f"{c}{R['check']}",
             f"={c}{R['assets']}-{c}{R['liab_eq']}", bold=True).fill = CHECK_OK
    _widths(ws, first=38, rest=15)
    return R


def _write_returns(ws: Worksheet, years: int, bs_rows: dict, is_rows: dict,
                   sensitivity: Optional[dict]):
    _title(ws, "Exit & Sponsor Returns",
           "Exit multiple applies to final-year EBITDA. IRR is a CAGR — valid because "
           "the sponsor has exactly one cash flow in and one out.")
    last = get_column_letter(3 + years - 1)
    R = {"ebitda": 5, "mult": 6, "ev": 7, "debt": 8, "cash": 9, "eq": 10,
         "spons": 12, "moic": 13, "irr": 14}
    for key, text, bold in (("ebitda", "Exit-year EBITDA", False),
                            ("mult", "Exit EV / EBITDA", False),
                            ("ev", "Exit enterprise value", True),
                            ("debt", "Less: debt outstanding", False),
                            ("cash", "Add: cash balance", False),
                            ("eq", "Exit equity value", True),
                            ("spons", "Sponsor equity invested", False),
                            ("moic", "MOIC", True), ("irr", "IRR", True)):
        _label(ws, f"A{R[key]}", text, bold=bold, indent=0 if bold else 1)

    _put(ws, f"B{R['ebitda']}", f"='{IS_}'!{last}{is_rows['ebitda']}")
    _put(ws, f"B{R['mult']}", f"={_a('exit')}", MULT)
    _put(ws, f"B{R['ev']}", f"=B{R['ebitda']}*B{R['mult']}", bold=True)
    _put(ws, f"B{R['debt']}", f"='{BS}'!{last}{bs_rows['debt']}")
    _put(ws, f"B{R['cash']}", f"='{BS}'!{last}{bs_rows['cash']}")
    _put(ws, f"B{R['eq']}", f"=B{R['ev']}-B{R['debt']}+B{R['cash']}", bold=True)
    _put(ws, f"B{R['spons']}", f"={_su('sponsor')}")
    _put(ws, f"B{R['moic']}", f"=B{R['eq']}/B{R['spons']}", MULT, bold=True)
    _put(ws, f"B{R['irr']}", f"=B{R['moic']}^(1/{_a('hold')})-1", PCT, bold=True)

    if sensitivity and sensitivity.get("irr"):
        start = 17
        ws.cell(row=start - 1, column=1, value="Sensitivity — IRR (exit multiple × leverage)"
                ).font = SECTION_FONT
        ws.cell(row=start + 1, column=1, value="Leverage ↓ / Exit →").font = Font(italic=True)
        for j, exit_m in enumerate(sensitivity["exit_multiples"]):
            c = ws.cell(row=start + 1, column=2 + j, value=exit_m)
            c.number_format, c.font, c.fill = MULT, HEADER_FONT, HEADER_FILL
        for i, lev in enumerate(sensitivity["leverage_multiples"]):
            c = ws.cell(row=start + 2 + i, column=1, value=lev)
            c.number_format, c.font = MULT, Font(bold=True)
            for j, val in enumerate(sensitivity["irr"][i]):
                cell = ws.cell(row=start + 2 + i, column=2 + j, value=val)
                cell.number_format = PCT
                cell.font = Font(color=INPUT_BLUE)
        note = ws.cell(row=start + 2 + len(sensitivity["leverage_multiples"]) + 1, column=1,
                       value="Grid values come from re-running the Python engine, not from "
                             "this workbook's formulas — hence blue. Excel Data Tables "
                             "cannot be written by openpyxl.")
        note.font = SUBTITLE_FONT
    _widths(ws, first=32, rest=14)


def _write_memo(ws: Worksheet, memo_text: Optional[str]):
    _title(ws, "Investment Memo", "Written by the Claude agent from the model outputs.")
    cell = ws["A4"]
    cell.value = memo_text or "(no memo)"
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 118


def build_workbook(company: CompanyFinancials, ltm_revenue: float, ltm_ebitda: float,
                   assumptions: LBOAssumptions, result: LBOResult,
                   sensitivity: Optional[dict] = None,
                   memo_text: Optional[str] = None) -> Workbook:
    """
    Build the live model. `result` is no longer written into the sheets — the workbook
    recomputes everything from the assumptions — but it stays in the signature because
    the caller has it and the recalculation test compares the two.
    """
    years = assumptions.hold_period_years
    wb = Workbook()
    wb.remove(wb.active)

    _write_assumptions(wb.create_sheet(ASM), company, ltm_revenue, ltm_ebitda, assumptions)
    _write_sources_uses(wb.create_sheet(SU))
    _write_debt_schedule(wb.create_sheet(DS), years)
    is_rows = _write_income_statement(wb.create_sheet(IS_), years)
    cf_rows = _write_cash_flow(wb.create_sheet(CF), years)
    bs_rows = _write_balance_sheet(wb.create_sheet(BS), years, cf_rows, is_rows)
    _write_returns(wb.create_sheet("Returns"), years, bs_rows, is_rows, sensitivity)
    _write_memo(wb.create_sheet("Investment Memo"), memo_text)

    # Tab colours: inputs, statements, outputs.
    wb[ASM].sheet_properties.tabColor = "0070C0"
    for name in (SU, DS, IS_, CF, BS):
        wb[name].sheet_properties.tabColor = "808080"
    wb["Returns"].sheet_properties.tabColor = "00B050"
    return wb


def save_workbook(wb: Workbook, path: str) -> str:
    wb.save(path)
    return path
