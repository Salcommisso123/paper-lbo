"""
Tests for the live Excel model.

Two levels:

  1. Offline (always runs) — structural invariants. No number outside the Assumptions
     sheet may be a hardcode, and the colour convention must be true of every cell.
  2. Recalculation (skipped without LibreOffice) — actually evaluates the formulas and
     checks the workbook reproduces the Python engine, and that the balance sheet ties.

(2) is the one that matters. It caught a real bug: the debt schedule's sweep pointed at
the cash-flow statement's NWC row instead of its free-cash-flow row, one line below.
With zero revenue growth NWC is zero, so the sweep silently became zero and the model
paid down only mandatory amortisation — every cell still looked plausible.
"""

import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from edgar import CompanyFinancials, FiscalYearFinancials  # noqa: E402
from excel_writer import (ASM, FORMULA_BLACK, INPUT_BLUE, LINK_GREEN,  # noqa: E402
                          build_workbook, save_workbook)
from lbo_engine import DebtTranche, LBOAssumptions, run_lbo  # noqa: E402

SOFFICE = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
LTM_REVENUE, LTM_EBITDA = 1_135_324_000.0, 101_106_000.0


def _fixture():
    """A deterministic company + assumptions pair, so the test never touches the network."""
    year = FiscalYearFinancials(fy=2025, revenue=LTM_REVENUE, ebitda=LTM_EBITDA,
                                cash=117_091_000.0, ppe_net=185_600_000.0,
                                current_assets=596_100_000.0,
                                current_liabilities=158_400_000.0)
    company = CompanyFinancials(ticker="SHOE", cik="0000895447",
                                company_name="SHOE STATION GROUP INC", years=[year])
    assumptions = LBOAssumptions(
        entry_ev_multiple=5.0, exit_ev_multiple=5.0, hold_period_years=5,
        leverage_multiple=4.5, revenue_growth_rate=0.0,
        da_pct_revenue=0.03, capex_pct_revenue=0.03, nwc_pct_of_revenue_change=0.15,
        tranches=[
            DebtTranche("Senior Term Loan", 0.75, 0.085, 0.01, priority=1),
            DebtTranche("Subordinated Notes", 0.25, 0.115, 0.0, priority=2),
        ])
    return company, assumptions, run_lbo(LTM_REVENUE, LTM_EBITDA, assumptions)


def _workbook(tmp_path: Path) -> Path:
    company, assumptions, result = _fixture()
    wb = build_workbook(company, LTM_REVENUE, LTM_EBITDA, assumptions, result,
                        memo_text="test memo")
    path = tmp_path / "model.xlsx"
    save_workbook(wb, str(path))
    return path


# ---- structural invariants (offline) -----------------------------------------

def test_only_the_assumptions_sheet_holds_hardcoded_numbers(tmp_path):
    """
    The whole point of the rewrite: the model must recompute, not report. Any bare
    number outside Assumptions (and the engine-produced sensitivity grid) is a cell
    that would not move when someone flexes an input.
    """
    wb = openpyxl.load_workbook(_workbook(tmp_path))
    offenders = []
    for sheet in wb:
        if sheet.title in (ASM, "Returns", "Investment Memo"):
            continue          # Assumptions is inputs; Returns holds the engine's grid
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    offenders.append(f"{sheet.title}!{cell.coordinate}={cell.value}")
    assert not offenders, f"hardcoded numbers outside Assumptions: {offenders[:8]}"


def test_colour_convention_matches_cell_contents(tmp_path):
    """Blue = hardcode, black = same-sheet formula, green = cross-sheet reference."""
    wb = openpyxl.load_workbook(_workbook(tmp_path))
    wrong = []
    for sheet in wb:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None or not cell.font or not cell.font.color:
                    continue
                rgb = str(cell.font.color.rgb or "")[-6:]
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    want = LINK_GREEN if "!" in cell.value else FORMULA_BLACK
                elif isinstance(cell.value, (int, float)):
                    want = INPUT_BLUE
                else:
                    continue
                if rgb != want:
                    wrong.append(f"{sheet.title}!{cell.coordinate} is {rgb}, expected {want}")
    assert not wrong, wrong[:8]


def test_three_statements_and_a_debt_schedule_are_present(tmp_path):
    wb = openpyxl.load_workbook(_workbook(tmp_path))
    for required in ("Assumptions", "Sources & Uses", "Debt Schedule",
                     "Income Statement", "Cash Flow", "Balance Sheet", "Returns"):
        assert required in wb.sheetnames


def test_no_formula_references_a_later_row_of_its_own_column_circularly(tmp_path):
    """Interest is on beginning balances, so nothing should need iterative calculation."""
    wb = openpyxl.load_workbook(_workbook(tmp_path))
    debt = wb["Debt Schedule"]
    interest = debt["C7"].value                      # senior interest, year 1
    assert interest.startswith("=C6*")               # beginning balance, not ending


# ---- recalculation (needs LibreOffice) ---------------------------------------

@pytest.mark.skipif(not Path(SOFFICE).exists(), reason="LibreOffice not installed")
def test_recalculated_workbook_reproduces_the_engine(tmp_path):
    source = _workbook(tmp_path)
    out = tmp_path / "recalc"
    subprocess.run([SOFFICE, "--headless", "--calc", "--convert-to",
                    "xlsx:Calc MS Excel 2007 XML", "--outdir", str(out), str(source)],
                   check=True, capture_output=True, timeout=180)

    wb = openpyxl.load_workbook(out / source.name, data_only=True)
    _, assumptions, engine = _fixture()
    last = chr(ord("C") + assumptions.hold_period_years - 1)

    def close(got, want, name):
        assert got is not None, f"{name}: no recalculated value"
        assert abs(got - want) <= max(1e-6, abs(want) * 1e-6), \
            f"{name}: workbook {got:,.4f} vs engine {want:,.4f}"

    close(wb["Sources & Uses"]["B13"].value, engine.sources_uses.sponsor_equity, "sponsor equity")
    close(wb["Cash Flow"]["C10"].value,
          engine.projections[0].levered_free_cash_flow_pre_sweep, "year 1 FCF")
    close(wb["Debt Schedule"]["C20"].value, engine.projections[0].cash_swept, "year 1 sweep")
    close(wb["Balance Sheet"][f"{last}13"].value,
          engine.projections[-1].total_debt_ending, "exit debt")
    close(wb["Balance Sheet"][f"{last}6"].value, engine.exit_cash_balance, "exit cash")
    close(wb["Returns"]["B10"].value, engine.exit_equity_value, "exit equity")
    close(wb["Returns"]["B13"].value, engine.sponsor_moic, "MOIC")
    close(wb["Returns"]["B14"].value, engine.sponsor_irr, "IRR")


@pytest.mark.skipif(not Path(SOFFICE).exists(), reason="LibreOffice not installed")
def test_balance_sheet_ties_every_year(tmp_path):
    source = _workbook(tmp_path)
    out = tmp_path / "recalc"
    subprocess.run([SOFFICE, "--headless", "--calc", "--convert-to",
                    "xlsx:Calc MS Excel 2007 XML", "--outdir", str(out), str(source)],
                   check=True, capture_output=True, timeout=180)
    bs = openpyxl.load_workbook(out / source.name, data_only=True)["Balance Sheet"]
    for column in "BCDEFG":                           # at close, then each of 5 years
        check = bs[f"{column}17"].value
        assert check is not None and abs(check) < 1.0, \
            f"balance sheet does not tie in column {column}: off by {check:,.2f}"
